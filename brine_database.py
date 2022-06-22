#!/usr/bin/env python

__author__ = 'Jonathan W. Driver, PhD'
__version__ = '1.3.1'
# Updates from 1.0.0: (1) sheet selector, (2) import of 2nd format, (3) Li+ and F-, (4) multi-source, (5) 1x, 2x, etc.,
# (6) selector clearing, (7) alphabetical ordering, (8) fix error if no brine imported, (9) automatic commit when
# sending to database
# Updates from 1.1.0: (1) additive tool for additive salts, e.g. Na2CO3, (2) PrintView to see salt composition
# Updates from 1.2.0: (1) NaCl a new additive salt, (2) UI update triggered when concentration or total mass adjusted
# Update from 1.3.0: Fixed bug related to submission of new brine due to NaCl additive from 1.3.0

from math import isnan, exp
from xlrd import open_workbook
from openpyxl import load_workbook
from dymo import print_to_dymo
from datetime import datetime
from PyQt5.QtWidgets import QWidget, QGridLayout, QDialog, QHBoxLayout, QVBoxLayout, QPushButton, QComboBox, QLabel, \
    QLineEdit, QListWidget, QMainWindow, QMessageBox, QFileDialog, QCheckBox
from PyQt5.QtCore import Qt, pyqtSignal
import db_tools
from db_tools import DBObjects

db_default_path = 'U:\\UEORLAB1\\Python Files\\'
# db_default_path = 'C:\\Users\\jdriver\\Dropbox\\Public\\PycharmProjects\\BrineDatabase\\'
# db_default_path = 'C:\\Users\\Jonathan\\Dropbox\\Public\\PycharmProjects\\BrineDatabase\\'
db_default_name = 'brine_database v1-3-0.db'

gui_width = 662 + 170
gui_height = 19 * 25 - 12


class BrineConnection(db_tools.DBConnection):

    def get_projects_for_company_name(self, c_name: str):

        params = {DBObjects.Params.CompName.value: c_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.SelectProjectsforCompanyName, params)

    def get_brines_for_project_name(self, p_name: str):

        params = {DBObjects.Params.ProjName.value: p_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.SelectBrinesforProjectName, params)

    def get_brine_data_by_name(self, b_name: str):

        params = {DBObjects.Params.BrineName.value: b_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.SelectBrineDatabyName, params)

    def exec_get_project_tags_for_brine(self, b_name: str):

        params = {DBObjects.Params.BrineName.value: b_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.GetProjectTagsBrine, params)

    def exec_add_new_brine(self, p_name: str, b_name: str, b_abbr: str, ion_ppms: list, add_salt_percs: list):

        p = DBObjects.Params
        params = {p.ProjName.value: p_name, p.BrineName.value: b_name, p.BrineAbbr.value: b_abbr,
                  p.Li.value: ion_ppms[0], p.Na.value: ion_ppms[1], p.K.value: ion_ppms[2], p.Mg.value: ion_ppms[3],
                  p.Ca.value: ion_ppms[4], p.Ba.value: ion_ppms[5], p.Sr.value: ion_ppms[6], p.Fe2.value: ion_ppms[7],
                  p.F.value: ion_ppms[8], p.Cl.value: ion_ppms[9], p.Br.value: ion_ppms[10],
                  p.HCO3.value: ion_ppms[11], p.SO4.value: ion_ppms[12], p.Na2CO3.value: add_salt_percs[0],
                  p.NaCl.value: add_salt_percs[1]}

        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.AddNewBrine, params)

    def exec_remove_brine(self, b_name: str):

        params = {DBObjects.Params.BrineName.value: b_name}
        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.RemoveBrine, params)

    def exec_update_brine(self, b_name: str, b_abbr: str, ion_ppms: list, add_salt_percs: list):

        p = DBObjects.Params
        params = {p.BrineName.value: b_name, p.BrineAbbr.value: b_abbr,
                  p.Li.value: ion_ppms[0], p.Na.value: ion_ppms[1], p.K.value: ion_ppms[2], p.Mg.value: ion_ppms[3],
                  p.Ca.value: ion_ppms[4], p.Ba.value: ion_ppms[5], p.Sr.value: ion_ppms[6], p.Fe2.value: ion_ppms[7],
                  p.F.value: ion_ppms[8], p.Cl.value: ion_ppms[9], p.Br.value: ion_ppms[10],
                  p.HCO3.value: ion_ppms[11], p.SO4.value: ion_ppms[12], p.Na2CO3.value: add_salt_percs[0],
                  p.NaCl.value: add_salt_percs[1]}

        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.UpdateBrine, params)

    def exec_add_project_tag_to_brine(self, b_name: str, p_name: str, alias: str):

        params = {DBObjects.Params.BrineName.value: b_name, DBObjects.Params.ProjName.value: p_name,
                  DBObjects.Params.BrineAlias.value: alias}
        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.AddProjectTagBrine, params)

    def exec_remove_project_tag_for_brine(self, b_name: str, p_name: str):

        params = {DBObjects.Params.BrineName.value: b_name, DBObjects.Params.ProjName.value: p_name}
        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.RemoveProjectTagBrine, params)

    def exec_get_brine_alias(self, b_name: str, p_name: str):

        params = {DBObjects.Params.BrineName.value: b_name, DBObjects.Params.ProjName.value: p_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.GetBrineAlias, params)

    def exec_update_brine_alias(self, alias: str, b_name: str, p_name: str):

        params = {DBObjects.Params.BrineAlias.value: alias, DBObjects.Params.BrineName.value: b_name,
                  DBObjects.Params.ProjName.value: p_name}
        return self.exec_stored_procedure_with_params_no_fetchall(DBObjects.Procedures.UpdateBrineAlias, params)

    def exec_search_brine_by_name_or_alias(self, b_name: str):

        params = {DBObjects.Params.BrineName.value: b_name}
        return self.exec_stored_procedure_with_params(DBObjects.Procedures.SearchBrineName, params)


class OptionDialog(QDialog):

    optionPicked = pyqtSignal(int)

    def __init__(self, parent, title: str, question: str, options: list, current: int=0):
        super(OptionDialog, self).__init__(parent=parent)
        try:
            print(parent.sf)
        except Exception as e:
            print(e)
        self.setWindowTitle(title)
        self.question = question
        self.setLayout(QVBoxLayout())
        self.result = -1

        self.listbox = QListWidget(self)
        self.listbox.doubleClicked.connect(self.set_option)
        self.listbox.addItems(options)
        self.listbox.setCurrentRow(current)
        self.layout().addWidget(self.listbox)

        self.optionPicked.connect(self.close_wrapper)

    def set_option(self, _):
        i = self.listbox.currentIndex().row()
        self.optionPicked.emit(i)

    def close_wrapper(self, _):
        self.close()


class Brine:

    ion_names = {"Li+": 'lithium', "Na+": 'sodium', "K+": 'potassium', "Mg++": 'magnesium', "Ca++": 'calcium',
                 "Ba++": 'barium', "Sr++": 'strontium', "Fe++": 'ironII', "F-": 'fluoride', "Cl-": 'chloride',
                 "Br-": 'bromide', "HCO3-": 'bicarbonate', "CO3--": 'carbonate', "SO4--": 'sulfate'}

    ion_weights = {"Li+": 6.94, "Na+": 22.99, "K+": 39.1, "Mg++": 24.305, "Ca++": 40.078, "Ba++": 137.327,
                   "Sr++": 87.62, "Fe++": 55.84, "F-": 19.0, "Cl-": 35.453, "Br-": 79.904, "HCO3-": 61.01,
                   "CO3--": 60.008, "SO4--": 96.06}

    def __init__(self, lithium=0., sodium=0., potassium=0., magnesium=0., calcium=0., barium=0.,
                 strontium=0., iron2=0., fluoride=0., chloride=0., bromide=0., bicarbonate=0., carbonate=0.,
                 sulfate=0.):

        self.composition = {"Li+": lithium, "Na+": sodium, "K+": potassium, "Mg++": magnesium, "Ca++": calcium,
                            "Ba++": barium, "Sr++": strontium, "Fe++": iron2, "F-": fluoride, "Cl-": chloride,
                            "Br-": bromide, "HCO3-": bicarbonate, "CO3--": carbonate, "SO4--": sulfate}
        self.name = ''

    def set_ion(self, ion, value):
        if ion in self.composition.keys():
            self.composition[ion] = value

    def add_ion(self, ion, value):
        if ion in self.composition.keys():
            self.composition[ion] += value

    def get_tds(self):
        tds = 0.
        for value in self.composition.values():
            tds += value
        return tds

    def get_viscosity(self, temp):

        # convert temp to Celsius, calculate TDS in ppk
        # temp = 5. * (temp - 32.) / 9.
        sal = 1.0e-3 * self.get_tds()

        # second-order model parameters for relative viscosity vs. temperature
        a = 1.474e-3 + 1.5e-5 * temp + 3.927e-8 * temp ** 2
        b = 1.0734e-5 - 8.5e-8 * temp + 2.23e-10 * temp ** 2
        mur = 1 + a * sal + b * sal ** 2
        # model of DI viscosity vs. temperature
        muw = exp(-3.79418 + 604.129 / (139.18 + temp))

        return muw * mur


class NominalBrine(Brine):

    source_salts = {"LiCl": 42.39, "NaCl": 58.443, "KCl": 74.553, "CaCl2_2H2O": 147.01, "MgCl2_6H2O": 203.31,
                    "BaCl2": 208.23, "SrCl2": 158.526, "FeCl2": 126.751, "NaF": 41.99, "NaHCO3": 84.007,
                    "Na2SO4": 142.04, "NaBr": 102.894, "SrCl2_6H2O": 266.64, "BaCl2_2H2O": 244.26}
    unique_sources = {"Li+": "LiCl", "K+": "KCl", "Ca++": "CaCl2_2H2O", "Mg++": "MgCl2_6H2O",
                      "Ba++": ["BaCl2", "BaCl2_2H2O"], "Fe++": "FeCl2", "Sr++": ["SrCl2", "SrCl2_6H2O"], "F-": "NaF",
                      "HCO3-": "NaHCO3", "SO4--": "Na2SO4", "Br-": "NaBr"}
    additive_salts = {"Na2CO3": [[2, 1], ["Na+", "CO3--"], 105.989], "NaCl": [[1, 1], ["Na+", "Cl-"], 58.443]}

    def __init__(self, lithium=0., sodium=0., potassium=0., magnesium=0., calcium=0., barium=0.,
                 strontium=0., iron2=0., fluoride=0., chloride=0., bromide=0., bicarbonate=0., carbonate=0.,
                 sulfate=0., sodium_bicarbonate=0., sodium_chloride=0.):

        Brine.__init__(self)
        self.nominal_composition = {"Li+": lithium, "Na+": sodium, "K+": potassium, "Mg++": magnesium, "Ca++": calcium,
                                    "Ba++": barium, "Sr++": strontium, "Fe++": iron2, "F-": fluoride, "Cl-": chloride,
                                    "Br-": bromide, "HCO3-": bicarbonate, "CO3--": carbonate, "SO4--": sulfate}
        self.add_salt_composition = {"Na2CO3": sodium_bicarbonate, "NaCl": sodium_chloride}
        self.salt_composition = {}
        self.source_options = {"Ba++": 1, "Sr++": 1}
        self.update_composition()

    def update_composition(self):
        self.salt_composition = {}
        ss = NominalBrine.source_salts
        for key, value in NominalBrine.unique_sources.items():
            if type(value) == list:
                value = value[self.source_options[key]]
            ion_ppm = self.nominal_composition[key]
            self.salt_composition[value] = ion_ppm * ss[value] / Brine.ion_weights[key]
        na_ppm = self.nominal_composition["Na+"]
        na_w = Brine.ion_weights["Na+"]

        self.salt_composition["NaCl"] = (na_ppm / na_w) * ss["NaCl"] - \
                                        (ss["NaCl"] / ss["NaHCO3"]) * self.salt_composition["NaHCO3"] - \
                                        (2. * ss["NaCl"] / ss["Na2SO4"]) * self.salt_composition["Na2SO4"] - \
                                        (ss["NaCl"] / ss["NaBr"]) * self.salt_composition["NaBr"] - \
                                        (ss["NaCl"] / ss["NaF"]) * self.salt_composition["NaF"]
        cl_ppm = 0.
        for key, value in self.salt_composition.items():
            if key.find("Cl") != -1:
                d = 1.
                if key.find("Cl2") != -1:
                    d = 2.
                cl_ppm += d * value * Brine.ion_weights["Cl-"] / NominalBrine.source_salts[key]
        self.set_ion("Cl-", cl_ppm)
        self.set_ion("Na+", na_ppm)
        self.set_ion("CO3--", 0)
        for key in self.unique_sources.keys():
            self.set_ion(key, self.nominal_composition[key])
        for key, value in self.add_salt_composition.items():
            info = NominalBrine.additive_salts[key]
            mm = info[2]
            for valency, ion in zip(info[0], info[1]):
                ion_mm = Brine.ion_weights[ion]
                ion_add_ppm = int((valency * ion_mm / mm) * value * 10000)
                self.add_ion(ion, ion_add_ppm)

    def set_add_salt_composition(self, add_salt: str, value: int):
        self.add_salt_composition[add_salt] = value

    def get_n_tds(self):
        n_tds = 0.
        for value in self.nominal_composition.values():
            n_tds += value
        return n_tds


class PositiveNumericEdit(QLineEdit):

    editingVetted = pyqtSignal()

    def __init__(self, parent, default_value):
        super(PositiveNumericEdit, self).__init__(parent=parent)
        self.default_value = default_value
        self.set_value_default()
        self.editingFinished.connect(self.check_value)

    def set_value_default(self):

        self.blockSignals(True)
        self.setText(str(self.default_value))
        self.blockSignals(False)

    def check_value(self):

        txt = self.text()

        try:
            val = float(txt)
            if val <= 0:
                raise ValueError

        except ValueError:

            self.set_value_default()

        self.editingVetted.emit()


class TableEntry(QLineEdit):

    editingVetted = pyqtSignal(int)

    def __init__(self, parent, row: int, col: int, data_type):
        super(TableEntry, self).__init__(parent=parent)
        self.row = row
        self.col = col
        self.data_type = data_type
        self.setAlignment(Qt.AlignCenter)
        self.textChanged.connect(self.highlight)
        self.editingFinished.connect(self.check_value)

    def highlight(self):

        ps = self.font().pointSize()
        self.setStyleSheet('QLineEdit {background: yellow};')
        f = self.font()
        f.setPointSize(ps)
        self.setFont(f)

    def un_highlight(self):

        ps = self.font().pointSize()
        self.setStyleSheet('QLineEdit {background: white};')
        f = self.font()
        f.setPointSize(ps)
        self.setFont(f)

    def check_value(self):

        txt = self.text()

        try:
            val = self.data_type(txt)
            if val < 0:
                raise ValueError

        except ValueError:
            self.setText('0')

        self.un_highlight()
        self.editingVetted.emit(self.row - 4)


class AskSheetOption(OptionDialog):

    def __init__(self, parent, sheet_names):
        super(AskSheetOption, self).__init__(parent, 'Select Sheet', '', sheet_names)
        self.parent = parent
        self.sheet_names = sheet_names


class AskSaltSourceOption(OptionDialog):

    def __init__(self, parent, ion, salt_names, current: int):
        super(AskSaltSourceOption, self).__init__(parent, 'Select Salt', '', salt_names, current)
        self.parent = parent
        self.salt_names = salt_names
        self.ion = ion


class IonLabel(QLabel):

    def __init__(self, parent, text: str, active: bool=False):
        super(IonLabel, self).__init__(parent=parent, text=text)
        self.active = active

    def run_func_and_reset_font(self, func):

        ps = self.font().pointSize()
        func()
        f = self.font()
        f.setPointSize(ps)
        self.setFont(f)

    def mouseDoubleClickEvent(self, a0):
        super(IonLabel, self).mouseDoubleClickEvent(a0)
        if not self.active:
            return

        ion = self.text()
        salt_names = self.parent().n_brine.unique_sources[ion]
        current = self.parent().n_brine.source_options[ion]
        dlg = AskSaltSourceOption(parent=self.parent(), ion=ion, salt_names=salt_names, current=current)
        dlg.optionPicked.connect(self.set_source)
        dlg.exec_()

    def set_source(self, i: int):

        self.parent().n_brine.source_options[self.text()] = i
        print(self.parent().n_brine.source_options)
        try:
            self.parent().n_brine.update_composition()
        except Exception as e:
            QMessageBox(parent=self.parent(), text=str(e)).exec_()

    def enterEvent(self, a0):
        super(IonLabel, self).enterEvent(a0)
        if not self.active:
            return

        self.run_func_and_reset_font(lambda: self.setStyleSheet('QLabel {color: blue};'))

    def leaveEvent(self, a0):
        super(IonLabel, self).leaveEvent(a0)
        if not self.active:
            return

        self.run_func_and_reset_font(lambda: self.setStyleSheet('QLabel {color: black};'))


class TagsManager(QDialog):

    def __init__(self, parent, cnxn: BrineConnection, brine_name: str):
        super(TagsManager, self).__init__(parent=parent)
        self.setFixedSize(500, 150)
        self.setWindowTitle('Tags for {}'.format(brine_name))
        g_lyt = QGridLayout()
        self.setLayout(g_lyt)

        self.cnxn = cnxn
        self.brine_name = brine_name
        self.brine_data = cnxn.get_brine_data_by_name(b_name=brine_name)[0]
        p_id = self.brine_data[1]
        project = cnxn.get_column_from_table_where(DBObjects.Tables.ProjectsTable, DBObjects.Columns.ProjectName,
                                                   DBObjects.Columns.ProjectID, p_id)[0][0]
        c_id = cnxn.get_column_from_table_where(DBObjects.Tables.ProjectsTable, DBObjects.Columns.CompanyID,
                                                DBObjects.Columns.ProjectID, p_id)[0][0]
        company = cnxn.get_column_from_table_where(DBObjects.Tables.CompaniesTable, DBObjects.Columns.CompanyName,
                                                   DBObjects.Columns.CompanyID, c_id)[0][0]

        self.main_project_label = QLabel(parent=self, text='Main Project: {} ({})'.format(project, company))
        self.main_project_name = project

        self.company_label = QLabel(parent=self, text='Company:')
        self.company_option = QComboBox(parent=self)
        self.company_option.currentTextChanged.connect(self.company_option_command)
        self.project_label = QLabel(parent=self, text='Project:')
        self.project_option = QComboBox(parent=self)
        self.project_option.currentTextChanged.connect(self.project_option_command)
        self.alias_label = QLabel(parent=self, text='Brine Alias:')
        self.alias_edit = QLineEdit(parent=self)
        self.tags_listbox = QListWidget(parent=self)
        self.tags_listbox.clicked.connect(lambda: self.remove_button.setEnabled(True))
        self.tags_listbox.clicked.connect(self.display_alias)
        self.remove_button = QPushButton(parent=self, text='Remove Tag')
        self.remove_button.clicked.connect(self.remove_button_command)
        self.add_button = QPushButton(parent=self, text='Add Tag')
        self.add_button.clicked.connect(self.add_button_command)

        self.add_project = ''
        self.project_choices = []

        g_lyt.addWidget(self.main_project_label, 0, 0, 1, 3)
        g_lyt.addWidget(self.tags_listbox, 1, 0, 2, 1)
        g_lyt.addWidget(self.company_label, 1, 1, 1, 1)
        g_lyt.addWidget(self.company_option, 1, 2, 1, 1)
        g_lyt.addWidget(self.project_label, 2, 1, 1, 1)
        g_lyt.addWidget(self.project_option, 2, 2, 1, 1)
        g_lyt.addWidget(self.alias_label, 3, 1, 1, 1)
        g_lyt.addWidget(self.alias_edit, 3, 2, 1, 1)
        g_lyt.addWidget(self.remove_button, 3, 0, 1, 1)
        g_lyt.addWidget(self.add_button, 4, 2, 1, 1)

        self.tags = []

        self.initialize_ui()
        self.company_option.setCurrentText(company)

    def block_signals_all(self, block: bool):

        self.company_option.blockSignals(block)
        self.project_option.blockSignals(block)

    def execute_func_with_block(self, func, *args):

        self.block_signals_all(True)
        func(*args)
        self.block_signals_all(False)

    def initialize_ui(self):

        self.company_option.clear()
        self.project_option.clear()
        self.add_button.setEnabled(False)
        self.remove_button.setEnabled(False)

        self.initialize_companies_option()
        self.initialize_tags_listbox()

    def initialize_companies_option(self):

        companies = self.cnxn.get_all_companies()
        company_list = [' ']
        for company in companies:
            company_list.append(company[0])

        self.company_option.addItems(company_list)
        # self.add_button.setEnabled(False)

    def initialize_projects_option(self, company: str):

        self.project_option.clear()
        self.project_choices = [' ']

        if company != ' ':
            projects = self.parent().cnxn.get_projects_for_company_name(company)
            for project in projects:
                self.project_choices.append(project[0])

        self.project_option.addItems(self.project_choices)
        self.add_button.setEnabled(False)

    def initialize_tags_listbox(self):

        self.tags_listbox.clear()
        tags = self.parent().cnxn.exec_get_project_tags_for_brine(self.brine_name)
        self.tags = []
        for tag in tags:
            self.tags.append(tag[0])

        self.tags_listbox.addItems(self.tags)

    def company_option_command(self, *args):

        company = args[0]
        self.execute_func_with_block(self.initialize_projects_option, company)

    def project_option_command(self, *args):

        self.add_project = args[0]
        if self.add_project != ' ' and self.add_project != self.main_project_name:
            self.add_button.setEnabled(True)
        else:
            self.add_button.setEnabled(False)

    def display_alias(self):

        p_name = self.tags_listbox.currentItem().text()
        try:
            self.alias_edit.clear()
            alias = self.parent().cnxn.exec_get_brine_alias(b_name=self.brine_name, p_name=p_name)[0][0]
            self.alias_edit.setText(alias)
        except Exception as e:
            QMessageBox(parent=self, text=str(e)).exec_()

    def add_button_command(self):

        self.parent().linked_main_window.run_login_with_functions(self.add_project_tag_to_brine,
                                                                  self.initialize_tags_listbox)

    def add_project_tag_to_brine(self):

        self.parent().cnxn.exec_add_project_tag_to_brine(self.brine_name, self.add_project, self.alias_edit.text())
        self.alias_edit.clear()

    def remove_button_command(self):

        self.parent().linked_main_window.run_login_with_functions(self.remove_project_tag_from_brine,
                                                                  self.initialize_tags_listbox)

    def remove_project_tag_from_brine(self):

        self.parent().cnxn.exec_remove_project_tag_for_brine(self.brine_name, self.tags_listbox.currentItem().text())

    def closeEvent(self, _):

        try:
            project = self.parent().brine_selector.project_option.currentText()
            self.parent().brine_selector.brine_listbox_refresh(project)
        except Exception as e:
            print(e)


class BrineView(QWidget):

    def __init__(self, parent, nb: NominalBrine):
        super(BrineView, self).__init__(parent=parent)
        self.brine_db_gui = parent
        self.parent = parent
        self.n_brine = nb
        self.m_or_k = 'k'
        self.import_type = '_'
        self.wb = None
        self.sheet_option = -1

        lyt = QGridLayout()
        self.setLayout(lyt)

        name_label = QLabel(parent=self, text='Brine Name:')
        self.name_entry = QLineEdit(self)
        lyt.addWidget(name_label, 0, 0, 1, 1)
        lyt.addWidget(self.name_entry, 0, 1, 1, 4)
        abbrev_label = QLabel(parent=self, text='Abbreviation:')
        self.abbrev_entry = QLineEdit(self)
        lyt.addWidget(abbrev_label, 1, 0, 1, 1)
        lyt.addWidget(self.abbrev_entry, 1, 1, 1, 4)
        alias_label = QLabel(parent=self, text='Brine Alias:')
        self.alias_entry = QLineEdit(parent=self)
        self.alias_entry.setEnabled(False)
        lyt.addWidget(alias_label, 2, 0, 1, 1)
        lyt.addWidget(self.alias_entry, 2, 1, 1, 4)

        self.headers = [QLabel(parent=self, text='Ion'), QLabel(parent=self, text='Target [ppm]'),
                        QLabel(parent=self, text='Actual [ppm]'), QLabel(parent=self, text='  Additive  '),
                        QLabel(parent=self, text='%')]
        j = 0
        for header in self.headers:
            lyt.addWidget(header, 3, j, 1, 1)
            j += 1

        i = 4
        self.labels = []
        self.entries = []
        self.actual_entries = []
        for key, value in nb.nominal_composition.items():
            label = IonLabel(parent=self, text=key)
            if key in NominalBrine.unique_sources.keys():
                if type(NominalBrine.unique_sources[key]) == list:
                    label.active = True

            lyt.addWidget(label, i, 0, 1, 1)
            self.labels.append(label)
            entry = TableEntry(self, i, 0, int)
            self.entries.append(entry)
            lyt.addWidget(entry, i, 1, 1, 1)
            if key == 'CO3--':
                entry.setEnabled(False)
            else:
                entry.editingVetted.connect(self.update_actual_entries)
            a_entry = QLineEdit(parent=self)
            a_entry.setAlignment(Qt.AlignCenter)
            a_entry.setEnabled(False)
            lyt.addWidget(a_entry, i, 2, 1, 1)
            self.actual_entries.append(a_entry)
            i += 1

        j = 4
        self.add_labels = []
        self.add_entries = []
        for key, value in nb.add_salt_composition.items():
            label = QLabel(parent=self, text=key)
            lyt.addWidget(label, j, 3, 1, 1)
            add_entry = TableEntry(self, j, 1, float)
            add_entry.editingVetted.connect(self.update_add_entries)
            add_entry.setAlignment(Qt.AlignCenter)
            self.add_labels.append(label)
            self.add_entries.append(add_entry)
            lyt.addWidget(add_entry, j, 4, 1, 1)
            j += 1

        self.tds_label = QLabel(parent=self, text='TDS')
        lyt.addWidget(self.tds_label, i, 0, 1, 1)
        self.tds_entry = QLineEdit(self)
        lyt.addWidget(self.tds_entry, i, 1, 1, 1)
        self.tds_entry.setText(str(int(nb.get_n_tds())))
        self.tds_entry.setAlignment(Qt.AlignCenter)
        self.tds_entry.setEnabled(False)
        self.actual_tds_entry = QLineEdit(self)
        lyt.addWidget(self.actual_tds_entry, i, 2, 1, 1)
        self.actual_tds_entry.setAlignment(Qt.AlignCenter)
        self.actual_tds_entry.setEnabled(False)

        self.button_frame = QWidget(parent=self)
        b_lyt = QHBoxLayout()
        self.button_frame.setLayout(b_lyt)
        self.import_button = QPushButton(parent=self.button_frame, text='Import')
        self.import_button.clicked.connect(self.import_command)
        self.import_button.setEnabled(False)
        self.send_button = QPushButton(parent=self.button_frame, text='Send')
        self.send_button.clicked.connect(self.send_command)
        self.send_button.setEnabled(False)

        b_lyt.addWidget(self.import_button)
        b_lyt.addWidget(self.send_button)
        lyt.addWidget(self.button_frame, i + 1, 0, 1, 3)

        self.print_view_button = QPushButton(parent=self, text='Print View')
        self.print_view_button.clicked.connect(self.print_view_command)
        self.print_view_button.setEnabled(False)
        lyt.addWidget(self.print_view_button, i + 1, 3, 1, 2)

        self.name_entry.setEnabled(False)
        self.enable_editables(False)

        # total_mass_label = tk.Label(self, text='Total Mass [g]:')
        # self.total_mass_entry = tk.Entry(self)
        # self.total_mass_entry_var = tk.StringVar(self)
        # self.total_mass_entry_var.set('4000')
        # self.total_mass_entry.config(textvariable=self.total_mass_entry_var, width=10)
        # total_mass_label.grid(row=i+2, column=3, pady=(10, 0))
        # self.total_mass_entry.grid(row=i+2, column=4, pady=(10, 0))
        # self.total_mass_entry.bind('<FocusOut>', self.create_and_display_print_list)
        # self.total_mass_entry.bind('<Return>', self.create_and_display_print_list)
        #
        # conc_factor_label = tk.Label(self, text='Conc. Factor:')
        # self.conc_factor_entry = tk.Entry(self)
        # self.conc_factor_entry_var = tk.StringVar(self)
        # self.conc_factor_entry_var.set('1')
        # self.conc_factor_entry.config(textvariable=self.conc_factor_entry_var, width=10)
        # conc_factor_label.grid(row=i+3, column=3, pady=(0, 10))
        # self.conc_factor_entry.grid(row=i+3, column=4, pady=(0, 10))
        # self.conc_factor_entry.bind('<FocusOut>', self.create_and_display_print_list)
        # self.conc_factor_entry.bind('<Return>', self.create_and_display_print_list)

        # self.print_button = tk.Button(self, text='Print', command=self.print_command)
        # self.print_button.config(width=8, height=2)
        # self.print_button.grid(row=i+2, column=3, rowspan=2, columnspan=2, pady=(15, 10))

        # self.config(highlightbackground='black', highlightthickness=1)
        # self.grid(row=0, column=1, rowspan=1, columnspan=2, padx=(5, 10), ipadx=5, pady=(5, 5))

    def enable_editables(self, enable: bool):

        self.abbrev_entry.setEnabled(enable)
        self.alias_entry.setEnabled(enable)

        for label, entry in zip(self.labels, self.entries):
            if label.text() != 'CO3--':
                entry.setEnabled(enable)

        for add_e in self.add_entries:
            add_e.setEnabled(enable)

    def update_actual_entries(self, i: int):

        txt = self.entries[i].text()
        if txt == '':
            txt = '0'
        ppm = int(txt)
        ion = self.labels[i].text()
        self.n_brine.nominal_composition[ion] = ppm
        self.n_brine.update_composition()

        for label, ae in zip(self.labels, self.actual_entries):
            ion = label.text()
            ae.setText(str(int(self.n_brine.composition[ion])))

        self.tds_entry.setText(str(int(self.n_brine.get_n_tds())))
        self.actual_tds_entry.setText(str(int(self.n_brine.get_tds())))

    def update_add_entries(self, i: int):

        perc = float(self.add_entries[i].text())
        salt = self.add_labels[i].text()
        self.n_brine.add_salt_composition[salt] = perc
        self.n_brine.update_composition()

        for j in range(len(self.entries)):
            self.update_actual_entries(j)

    def sheet_option_picked(self, i: int):

        self.sheet_option = i

    def wire_send_to_update(self):

        self.send_button.disconnect()
        self.send_button.clicked.connect(self.update_command)

    def browse_brine(self, brine_data: tuple):

        self.send_button.setEnabled(False)
        self.import_button.setEnabled(False)

        self.name_entry.clear()
        self.name_entry.setText(brine_data[2])
        self.abbrev_entry.clear()
        self.abbrev_entry.setText(brine_data[3])

        ion_names_list = ['lithium', 'sodium', 'potassium', 'magnesium', 'calcium', 'barium', 'strontium', 'ironII',
                          'fluoride', 'chloride', 'bromide', 'bicarbonate', 'carbonate', 'sulfate']

        for j, label in enumerate(self.labels):
            ion = label.text()
            entry = self.entries[j]
            if ion != 'CO3--':
                entry.clear()
                ind = ion_names_list.index(Brine.ion_names[ion])
                entry.setText(str(brine_data[ind + 4]))
                entry.editingFinished.emit()

        add_salt_list = ['Na2CO3', 'NaCl']

        for j, label in enumerate(self.add_labels):
            add_salt = label.text()
            entry = self.add_entries[j]
            entry.clear()
            ind = add_salt_list.index(add_salt)
            entry.setText(str(brine_data[ind + 18]))
            entry.editingFinished.emit()

        self.send_button.disconnect()
        self.send_button.clicked.connect(self.send_command)
        self.name_entry.setEnabled(False)
        self.enable_editables(False)
        self.print_view_button.setEnabled(True)

        project = self.brine_db_gui.brine_selector.project_option.currentText()
        alias = self.brine_db_gui.cnxn.exec_get_brine_alias(brine_data[2], project)[0][0]
        self.alias_entry.setText(alias)

    def import_command(self):

        drive = self.brine_db_gui.linked_main_window.u_or_q
        f_name = QFileDialog.getOpenFileName(parent=self, caption='Get Import File...',
                                             directory=drive + ':\\UEORLAB1\\',
                                             filter='*.xls*')
        file_name = f_name[0]
        ext = '.' + file_name.split('.')[1]

        def ask_which_sheet(parent: BrineView, sheet_names: list, get_sheet_func):
            if len(sheet_names) == 1:
                self.read_and_load_data(get_sheet_func(sheet_names[0]))
            else:
                option_dialog = AskSheetOption(parent, sheet_names)
                option_dialog.optionPicked.connect(self.sheet_option_picked)
                option_dialog.exec_()
                parent.read_and_load_data(get_sheet_func(sheet_names[parent.sheet_option]))

        if ext == '.xls':
            self.import_type = ext
            wb = open_workbook(file_name)
            self.wb = wb
            ask_which_sheet(self, wb.sheet_names(), wb.sheet_by_name)
        elif ext == '.xlsx':
            self.import_type = ext
            wb = load_workbook(file_name)
            self.wb = wb
            ask_which_sheet(self, wb.get_sheet_names(), wb.get_sheet_by_name)
        else:
            return

    def read_and_load_data(self, ws):

        data_values = [[], [], [], [], [], []]
        if self.import_type == '_':
            return
        elif self.import_type == '.xls':
            for i in range(20):
                for ii in range(6):
                    data_values[ii].append(ws.cell_value(i, ii))
        else:
            for i in range(20):
                for ii in range(6):
                    data_values[ii].append(ws.cell(row=i+1, column=ii+1).value)
        j = -1
        jj = -1
        for k in range(10):
            if data_values[0][k] == 'Ion':
                j = k
            elif data_values[0][k] == 'Cation':
                jj = k

        if j != -1:
            i = j + 1
            while isinstance(data_values[0][i], str) and data_values[0][i]:
                ion = data_values[0][i]
                for j, label in enumerate(self.labels):
                    if label.text() == ion or (label.text() == 'SO4--' and ion == 'SO4='):
                        entry = self.entries[j]
                        entry.clear()
                        if data_values[1][i] is not None and type(data_values[1][i]) is not str \
                                and not isnan(data_values[1][i]):
                            entry.setText(str(int(data_values[1][i])))
                        else:
                            entry.setText(str(0))
                        entry.editingFinished.emit()

                i += 1
        elif jj != -1:
            ion_name_list = list(Brine.ion_names.values())
            ion_list = list(Brine.ion_names.keys())
            cols = [0, 3]
            for col in cols:
                i = jj + 1
                while isinstance(data_values[col][i], str) and data_values[col][i]:
                    ion_name = data_values[col][i].lower()
                    ion_name = ion_name.strip()
                    if ion_name == 'flouride':
                        ion_name = 'fluoride'
                    ind = ion_name_list.index(ion_name)
                    ion = ion_list[ind]
                    for j, label in enumerate(self.labels):
                        if label.text() == ion or (label.text() == 'SO4--' and ion == 'SO4='):
                            entry = self.entries[j]
                            entry.clear()
                            if data_values[col+2][i] is not None and type(data_values[col+2][i]) is not str \
                                    and not isnan(data_values[col+2][i]):
                                entry.setText(str(int(data_values[col+2][i])))
                            else:
                                entry.setText(str(0))
                            entry.editingFinished.emit()
                    i += 1

    def get_brine_info_and_data(self):

        name = self.name_entry.text()
        if not name.strip():
            QMessageBox(parent=self, text='Cannot send brine: Name is empty.').exec_()
            return

        abbr = self.abbrev_entry.text()
        if not abbr.strip():
            QMessageBox(parent=self, text='Cannot send brine: Abbreviation is empty.').exec_()
            return

        ion_names_list = ['lithium', 'sodium', 'potassium', 'magnesium', 'calcium', 'barium', 'strontium', 'ironII',
                          'fluoride', 'chloride', 'bromide', 'bicarbonate', 'sulfate']

        values = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for key, value in self.n_brine.nominal_composition.items():
            if key != 'CO3--':
                ion = Brine.ion_names[key]
                ind = ion_names_list.index(ion)
                values[ind] = value

        add_names_list = ['Na2CO3', 'NaCl']
        add_values = [0., 0.]
        for key, value in self.n_brine.add_salt_composition.items():
            ind = add_names_list.index(key)
            add_values[ind] = float(value)

        return name, abbr, values, add_values

    def send_command(self):

        self.brine_db_gui.linked_main_window.run_login_with_functions(self.send_brine, None)

    def send_brine(self):

        p_name = self.brine_db_gui.brine_selector.project_option.currentText()

        name, abbr, values, add_values = self.get_brine_info_and_data()

        self.brine_db_gui.cnxn.exec_add_new_brine(p_name=p_name, b_name=name, b_abbr=abbr, ion_ppms=values,
                                                  add_salt_percs=add_values)

        self.import_button.setEnabled(False)
        self.send_button.setEnabled(False)
        self.brine_db_gui.brine_selector.brine_listbox_refresh(p_name)

    def update_command(self):

        self.brine_db_gui.linked_main_window.run_login_with_functions(self.update_brine, None)

    def update_brine(self):

        name, abbr, values, add_values = self.get_brine_info_and_data()
        alias = self.alias_entry.text()
        project = self.brine_db_gui.brine_selector.project_option.currentText()

        self.brine_db_gui.cnxn.exec_update_brine(b_name=name, b_abbr=abbr, ion_ppms=values, add_salt_percs=add_values)
        self.brine_db_gui.cnxn.exec_update_brine_alias(alias=alias, b_name=name, p_name=project)

        self.import_button.setEnabled(False)
        self.send_button.setEnabled(False)

    def print_view_command(self):

        PrintView(parent=self).exec_()


class BrineSearchEdit(QLineEdit):

    search_brine = pyqtSignal(str)

    def __init__(self, parent):
        super(BrineSearchEdit, self).__init__(parent=parent)

    def keyPressEvent(self, a0):
        super(BrineSearchEdit, self).keyPressEvent(a0)
        if a0.key() in (Qt.Key_Enter, Qt.Key_Return) and self.text():
            self.search_brine.emit(self.text())


class BrineSelector(QWidget):

    def __init__(self, parent: QWidget):
        super(BrineSelector, self).__init__(parent=parent)
        self.brine_db_gui = parent
        lyt = QGridLayout()
        self.setLayout(lyt)

        name_alias_widget = QWidget(self)
        name_alias_widget.setLayout(QHBoxLayout())
        self.name_alias_label = QLabel(parent=self, text='Search by Name or Alias:')
        self.name_alias_edit = BrineSearchEdit(parent=self)
        self.name_alias_edit.search_brine.connect(self.search_for_brine)
        name_alias_widget.layout().addWidget(self.name_alias_label)
        name_alias_widget.layout().addWidget(self.name_alias_edit)
        name_alias_widget.layout().setContentsMargins(0, 0, 0, 0)

        self.company_label = QLabel(parent=self, text='Company:')
        self.company_option = QComboBox(parent=self)
        self.company_option.currentTextChanged.connect(self.company_option_command)
        self.project_label = QLabel(parent=self, text='Project:')
        self.project_option = QComboBox(parent=self)
        self.project_option.currentTextChanged.connect(self.project_option_command)

        self.brine_label = QLabel(parent=self, text='Brine List:')
        self.brine_listbox = QListWidget(parent=self)
        self.brine_listbox.doubleClicked.connect(self.browse_brine_command)
        self.brine_listbox.clicked.connect(lambda: self.enable_edit_remove(True))
        self.brine_listbox.doubleClicked.connect(lambda: self.enable_edit_remove(True))

        aer_widget = QWidget(self)
        aer_widget.setLayout(QHBoxLayout())
        self.add_button = QPushButton(parent=self, text='Add')
        self.add_button.clicked.connect(self.add_brine_command)
        self.edit_button = QPushButton(parent=self, text='Edit')
        self.edit_button.clicked.connect(self.edit_brine_command)
        self.tags_button = QPushButton(parent=self, text='Tags')
        self.tags_button.clicked.connect(self.tags_command)
        self.remove_button = QPushButton(parent=self, text='Remove')
        self.remove_button.clicked.connect(self.remove_brine_command)
        aer_widget.layout().addWidget(self.add_button)
        aer_widget.layout().addWidget(self.edit_button)
        aer_widget.layout().addWidget(self.tags_button)
        aer_widget.layout().addWidget(self.remove_button)

        lyt.addWidget(name_alias_widget, 0, 0, 1, 2)
        lyt.addWidget(self.company_label, 1, 0, 1, 1)
        lyt.addWidget(self.company_option, 1, 1, 1, 1)
        lyt.addWidget(self.project_label, 2, 0, 1, 1)
        lyt.addWidget(self.project_option, 2, 1, 1, 1)
        lyt.addWidget(self.brine_label, 3, 0, 1, 1)
        lyt.addWidget(self.brine_listbox, 4, 0, 1, 2)
        lyt.addWidget(aer_widget, 5, 0, 1, 2)

        self.brine_list = []
        self.project_choices = [' ']

        self.execute_func_with_block(self.initialize_ui)

    def block_signals_all(self, block: bool):

        self.company_option.blockSignals(block)
        self.project_option.blockSignals(block)
        self.brine_listbox.blockSignals(block)

    def enable_edit_remove(self, enable: bool):

        self.edit_button.setEnabled(enable)
        self.tags_button.setEnabled(enable)
        self.remove_button.setEnabled(enable)

    def execute_func_with_block(self, func, *args):

        self.block_signals_all(True)
        func(*args)
        self.block_signals_all(False)

    def initialize_ui(self):

        self.company_option.clear()
        self.project_option.clear()
        self.brine_listbox.clear()

        self.initialize_companies_option()

    def initialize_companies_option(self):

        companies = self.brine_db_gui.cnxn.get_all_companies()
        company_list = [' ']
        for company in companies:
            company_list.append(company[0])

        self.company_option.addItems(company_list)
        self.enable_edit_remove(False)
        self.add_button.setEnabled(False)

    def initialize_projects_option(self, company: str):

        self.project_option.clear()
        self.project_choices = [' ']
        if company != ' ':
            projects = self.brine_db_gui.cnxn.get_projects_for_company_name(company)
            for project in projects:
                self.project_choices.append(project[0])

        self.project_option.addItems(self.project_choices)
        self.enable_edit_remove(False)
        self.add_button.setEnabled(False)

    def company_option_command(self, *args):

        company = args[0]
        self.execute_func_with_block(self.initialize_projects_option, company)

    def project_option_command(self, *args):

        project = args[0]

        if project == ' ':
            self.execute_func_with_block(self.brine_listbox.clear)
            self.brine_listbox.setEnabled(False)
        else:
            self.brine_listbox.setEnabled(True)
            self.execute_func_with_block(lambda: self.brine_listbox_refresh(project))

    def brine_listbox_refresh(self, project: str):

        self.brine_listbox.clear()

        try:
            brines = self.brine_db_gui.cnxn.get_brines_for_project_name(project)

        except Exception as e:
            QMessageBox(parent=self.brine_db_gui, text=str(e)).exec_()
            return

        self.brine_list = brines

        for i in range(len(brines)):
            self.brine_list[i] = brines[i][0]

        self.brine_listbox.addItems(self.brine_list)
        self.enable_edit_remove(False)
        self.add_button.setEnabled(True)

    def search_for_brine(self, brine_name: str):

        results = self.brine_db_gui.cnxn.exec_search_brine_by_name_or_alias(brine_name)
        if results:
            try:
                BrineSearchSelector(parent=self, search_results=results).exec_()
            except Exception as e:
                print(e)

    def browse_brine_command(self, _):

        brine = self.brine_listbox.currentItem().text()

        brine_data_temp = self.brine_db_gui.cnxn.get_brine_data_by_name(brine)[0]
        brine_data = (*brine_data_temp[:-3], '--', *brine_data_temp[-3:],)

        try:
            self.brine_db_gui.brine_view.browse_brine(brine_data=brine_data)
        except Exception as e:
            QMessageBox(parent=self, text=str(e)).exec_()

    def add_brine_command(self):

        brine_data = (0, 0, '', '', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, '--', 0, 0., 0.,)
        self.brine_db_gui.brine_view.browse_brine(brine_data=brine_data)
        self.brine_db_gui.brine_view.send_button.setEnabled(True)
        self.brine_db_gui.brine_view.import_button.setEnabled(True)
        self.brine_db_gui.brine_view.print_view_button.setEnabled(False)
        self.enable_edit_remove(False)
        self.brine_db_gui.brine_view.name_entry.setEnabled(True)
        self.brine_db_gui.brine_view.enable_editables(True)
        self.brine_db_gui.brine_view.alias_entry.setEnabled(False)

    def edit_brine_command(self):

        self.brine_listbox.doubleClicked.emit(self.brine_listbox.currentIndex())
        self.brine_db_gui.brine_view.wire_send_to_update()
        self.brine_db_gui.brine_view.send_button.setEnabled(True)
        self.brine_db_gui.brine_view.enable_editables(True)
        self.brine_db_gui.brine_view.print_view_button.setEnabled(True)

    def tags_command(self):

        brine = self.brine_listbox.currentItem().text()
        TagsManager(self.brine_db_gui, self.brine_db_gui.cnxn, brine).exec_()

    def remove_brine_command(self):

        self.brine_db_gui.linked_main_window.run_login_with_functions(self.remove_brine, None)

    def remove_brine(self):

        brine = self.brine_listbox.currentItem().text()
        project = self.project_option.currentText()
        self.brine_db_gui.cnxn.exec_remove_brine(brine)
        self.brine_listbox_refresh(project)


class BrineSearchSelector(QDialog):

    def __init__(self, parent: BrineSelector, search_results: list):
        super(BrineSearchSelector, self).__init__(parent=parent)
        self.parent = parent
        self.setWindowTitle('Search Results')
        sf = parent.brine_db_gui.linked_main_window.sf
        self.setFixedSize(int(sf * 600), int(sf * 300))
        self.search_results = search_results
        self.setLayout(QVBoxLayout())

        self.results_list = QListWidget(parent=self)
        to_add = list()
        for result in search_results:
            if result[2] is None:
                to_add.append('{}, {}, Company = {}, Project = {}'.format(result[0], result[1], result[3], result[4]))
            else:
                to_add.append('{}, {} ({}), Company = {}, Project = {}'.format(*result))

        self.results_list.addItems(to_add)
        self.layout().addWidget(self.results_list)
        self.results_list.itemDoubleClicked.connect(self.set_brine_selector_to_brine)

    def set_brine_selector_to_brine(self, _):

        i = self.results_list.currentIndex().row()
        brine_info = self.search_results[i]
        company_index = self.parent.company_option.findText(brine_info[3])
        self.parent.company_option.setCurrentIndex(company_index)
        project_index = self.parent.project_option.findText(brine_info[4])
        self.parent.project_option.setCurrentIndex(project_index)
        item = self.parent.brine_listbox.findItems(brine_info[1], Qt.MatchExactly)[0]
        self.parent.brine_listbox.setCurrentItem(item)

        self.close()


class PrintView(QDialog):

    def __init__(self, parent: BrineView):
        super(PrintView, self).__init__(parent=parent)
        self.setWindowTitle('Print Label: {}'.format(parent.abbrev_entry.text()))
        self.setFixedSize(250, 550)
        lyt = QGridLayout()
        self.setLayout(lyt)

        self.print_list_labels = []
        for i in range(18):
            label = QLabel(parent=self)
            font = label.font()
            font.setFamily('Courier')
            self.print_list_labels.append(label)
            lyt.addWidget(label, i, 0, 1, 2)
            if i == 0:
                label.setAlignment(Qt.AlignCenter)
                font.setBold(True)

            label.setFont(font)

        total_mass_label = QLabel(parent=self, text='Total Mass [g]:')
        conc_label = QLabel(parent=self, text='Conc. Factor (X):')
        self.total_mass_edit = PositiveNumericEdit(parent=self, default_value=4000)
        self.total_mass_edit.editingVetted.connect(self.create_and_display_print_list)
        self.conc_edit = PositiveNumericEdit(parent=self, default_value=1.0)
        self.conc_edit.editingVetted.connect(self.create_and_display_print_list)
        self.print_ion_label_checkbox = QCheckBox(parent=self, text='Print Ion Label?')
        self.print_ion_label_checkbox.setChecked(False)
        self.print_button = QPushButton(parent=self, text='Print')
        self.print_button.clicked.connect(self.print_command)

        lyt.addWidget(self.print_ion_label_checkbox, 18, 0, 1, 2)
        lyt.addWidget(total_mass_label, 19, 0, 1, 1)
        lyt.addWidget(self.total_mass_edit, 19, 1, 1, 1)
        lyt.addWidget(conc_label, 20, 0, 1, 1)
        lyt.addWidget(self.conc_edit, 20, 1, 1, 1)
        lyt.addWidget(self.print_button, 21, 0, 1, 2)

        self.create_and_display_print_list()

    def create_and_display_print_list(self):

        sorted_print_list = self.create_print_list()

        if sorted_print_list:
            self.display_print_list(sorted_print_list)

    def create_print_list(self) -> list:

        name = self.parent().abbrev_entry.text()
        conc_factor = float(self.conc_edit.text())
        total_mass = float(self.total_mass_edit.text())
        water_mass = total_mass

        print_list = [name]

        masses = []
        for key, value in self.parent().n_brine.salt_composition.items():
            if key in self.parent().n_brine.add_salt_composition.keys():
                value += 10000. * self.parent().n_brine.add_salt_composition[key]
            grams = 1e-6 * total_mass * value * conc_factor
            water_mass -= grams
            if grams > 0:
                masses.append(grams)
                print_key = key
                if print_key.find('H2O') != -1:
                    print_key = print_key[:-3] + 'W'
                print_list.append('{:>7.3f}'.format(grams) + ' g ' + print_key)

        for key, value in self.parent().n_brine.add_salt_composition.items():
            if key in self.parent().n_brine.salt_composition.keys():
                continue
            grams = 1e-2 * total_mass * value * conc_factor
            water_mass -= grams
            if grams > 0:
                masses.append(grams)
                print_key = key
                if print_key.find('H2O') != -1:
                    print_key = print_key[:-3] + 'W'
                print_list.append('{:>7.3f}'.format(grams) + ' g ' + print_key)

        s_masses = sorted(masses, reverse=True)
        sorted_print_list = [print_list[0]]
        for mass in s_masses:
            ind = masses.index(mass)
            sorted_print_list.append(print_list[ind + 1])
        sorted_print_list.append('{:>7.2f}'.format(water_mass) + ' g ' + 'dH2O')

        return sorted_print_list

    def create_ion_print_list(self) -> list:

        name = self.parent().abbrev_entry.text()
        conc_factor = float(self.conc_edit.text())
        composition = self.parent().n_brine.composition

        print_list = [name]
        skipped_cations = 0
        skipped_anions = 0
        counter = 0
        for ion, ppm in composition.items():

            counter += 1

            is_cation = False
            if ion.find('+') != -1:
                is_cation = True

            if not is_cation and counter == 9:
                for i in range(skipped_cations):
                    print_list.append('')

            if ppm == 0 and is_cation:
                skipped_cations += 1
                continue
            if ppm == 0 and not is_cation:
                skipped_anions += 1
                continue

            while len(ion) < 5:
                ion += ' '
            print_list.append('{}: {:6d} ppm'.format(ion, int(ppm * conc_factor)))

        for i in range(skipped_anions):
            print_list.append('')

        print_list.append('TDS  : {:6d} ppm'.format(int(self.parent().n_brine.get_tds())))

        return print_list

    def display_print_list(self, print_list: list):

        for label in self.print_list_labels:
            label.setText('')

        for i, item in enumerate(print_list):
            self.print_list_labels[i].setText(item)

    def print_command(self):

        conc_factor = float(self.conc_edit.text())
        sorted_print_list = self.create_print_list()

        if not sorted_print_list:
            return

        if len(sorted_print_list) < 14:
            for i in range(14 - len(sorted_print_list)):
                sorted_print_list.append('')
            label_file = 'my.label'
        elif len(sorted_print_list) == 14:
            label_file = 'my.label'
        elif 14 < len(sorted_print_list) < 17:
            for i in range(17 - len(sorted_print_list)):
                sorted_print_list.append('')
            label_file = 'brine_with_additives.label'
        elif len(sorted_print_list) > 17:
            print('print list is too long: ', len(sorted_print_list))
            label_file = ''
        else:
            label_file = 'brine_with_additives.label'

        sorted_print_list.append('DATE: ' + datetime.today().strftime('%Y-%m-%d'))
        if conc_factor != 1.:
            sorted_print_list[0] = str(conc_factor) + 'x' + sorted_print_list[0]
        # print(len(sorted_print_list), label_file, sorted_print_list)
        print_to_dymo(sorted_print_list, label_file)

        if self.print_ion_label_checkbox.isChecked():
            ion_print_list = self.create_ion_print_list()
            # print(ion_print_list)
            print_to_dymo(ion_print_list, 'ion_comp.label')

        self.close()


class BrineDatabaseGUI(QWidget):

    def __init__(self, cnxn: db_tools.pyodbc.Connection, linked_main_window: QMainWindow):
        super(BrineDatabaseGUI, self).__init__()
        self.cnxn = None
        self.set_connection(cnxn=cnxn)
        self.setLayout(QHBoxLayout())

        self.linked_main_window = linked_main_window
        # path = db_default_path
        self.brine_selector = BrineSelector(self)
        self.brine_view = BrineView(self, NominalBrine())
        self.layout().addWidget(self.brine_selector)
        self.layout().addWidget(self.brine_view)

        # self.print_view = PrintView(self)

    def set_connection(self, cnxn: db_tools.pyodbc.Connection):

        self.cnxn = BrineConnection(cnxn)
